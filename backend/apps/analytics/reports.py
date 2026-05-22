"""
Report generators for PDF (ReportLab) and Excel (openpyxl) exports.
Both functions return a seeked BytesIO buffer ready to stream as an HTTP response.
"""

import datetime
import io

from django.db.models import Count, Q, Sum

from apps.projects.models import Project
from apps.tasks.models import Task
from apps.users.models import Client
from apps.feedback.models import Feedback
from apps.timelog.models import TimeLog


def _weighted_designer_rate(log_qs, actual_hours):
    """Return the logged-hours weighted designer rate, imputing missing rates from the known average."""
    if actual_hours <= 0:
        return None

    weighted_total = 0.0
    rated_hours = 0.0
    for row in log_qs.values('designer__hourly_rate').annotate(logged_hours=Sum('hours_spent')):
        hourly_rate = row['designer__hourly_rate']
        logged_hours = float(row['logged_hours'] or 0)
        if hourly_rate is None:
            continue
        weighted_total += float(hourly_rate) * logged_hours
        rated_hours += logged_hours

    # Impute missing rates from known average (matches ProfitMarginView logic)
    if actual_hours > 0 and rated_hours > 0:
        known_avg = weighted_total / rated_hours
        imputed = known_avg * (actual_hours - rated_hours)
        return (weighted_total + imputed) / actual_hours
    return None


# ─────────────────────────────────────────────────────────────────────────────
# PDF — project profitability summary
# ─────────────────────────────────────────────────────────────────────────────

def generate_project_pdf(project_id: int) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame,
        Table, TableStyle, Paragraph, Spacer,
        HRFlowable,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT

    # ── Data ──────────────────────────────────────────────────────────────────
    project = Project.objects.select_related('client__user').get(id=project_id)

    task_qs = Task.objects.filter(project=project)
    task_agg = task_qs.aggregate(
        estimated=Sum('estimated_hours'),
        total=Count('id', distinct=True),
        completed=Count('id', filter=Q(status='Completed'), distinct=True),
        unplanned=Count('id', filter=Q(is_unplanned=True), distinct=True),
    )

    # Query actual hours separately to avoid join duplication inflating estimates
    actual_hours = float(
        TimeLog.objects.filter(task__project=project).aggregate(t=Sum('hours_spent'))['t'] or 0
    )
    estimated_h     = float(task_agg['estimated'] or 0)
    budget_hours    = float(project.budget_hours  or 0)
    budget_amount   = float(project.budget_amount or 0)
    total_tasks     = task_agg['total']     or 0
    completed_tasks = task_agg['completed'] or 0
    unplanned_tasks = task_agg['unplanned'] or 0

    ehr          = budget_amount / actual_hours if actual_hours > 0 else 0
    budget_util  = actual_hours / budget_hours * 100 if budget_hours > 0 else 0
    scope_creep  = unplanned_tasks / total_tasks * 100 if total_tasks > 0 else 0
    task_done_pct = completed_tasks / total_tasks * 100 if total_tasks > 0 else 0

    fb_agg = Feedback.objects.filter(project=project).aggregate(
        revisions=Count('id', filter=Q(category='Revision')),
        approvals=Count('id', filter=Q(category='Approval')),
    )
    revisions = fb_agg['revisions'] or 0
    approvals  = fb_agg['approvals'] or 0

    weighted_rate = _weighted_designer_rate(TimeLog.objects.filter(task__project=project), actual_hours)
    margin = (
        (ehr - weighted_rate) / ehr * 100
        if weighted_rate is not None and ehr > 0 else None
    )

    # --- Projected metrics for incomplete projects (matches dashboard) ---
    is_completed = project.status == 'Completed'
    projected_ehr = None
    projected_margin = None
    target_ehr = budget_amount / budget_hours if budget_hours > 0 else None

    if not is_completed and actual_hours > 0 and total_tasks > 0 and completed_tasks > 0:
        completion_ratio = completed_tasks / total_tasks
        projected_hours = actual_hours / completion_ratio
        projected_ehr = budget_amount / projected_hours
        if weighted_rate is not None and projected_ehr > 0:
            projected_margin = (projected_ehr - weighted_rate) / projected_ehr * 100
    elif not is_completed and target_ehr is not None and weighted_rate is not None:
        # No completed tasks yet — fall back to margin at budget hours
        projected_margin = (target_ehr - weighted_rate) / target_ehr * 100

    # Decide what to display
    display_ehr = ehr if is_completed else (projected_ehr if projected_ehr is not None else ehr)
    display_margin = margin if is_completed else (projected_margin if projected_margin is not None else margin)
    ehr_label = 'EFF. HOURLY RATE' if is_completed else 'PROJ. HOURLY RATE'
    margin_label = 'Profit Margin' if is_completed else 'Projected Margin'

    # ── Palette ───────────────────────────────────────────────────────────────
    C_PRIMARY    = colors.HexColor('#6366f1')
    C_DARK       = colors.HexColor('#0e0d28')
    C_SLATE_900  = colors.HexColor('#0f172a')
    C_SLATE_700  = colors.HexColor('#334155')
    C_SLATE_600  = colors.HexColor('#475569')
    C_SLATE_400  = colors.HexColor('#94a3b8')
    C_SLATE_200  = colors.HexColor('#e2e8f0')
    C_SLATE_100  = colors.HexColor('#f1f5f9')
    C_SLATE_50   = colors.HexColor('#f8fafc')
    C_SUCCESS    = colors.HexColor('#10b981')
    C_WARNING    = colors.HexColor('#f59e0b')
    C_DANGER     = colors.HexColor('#ef4444')
    C_WHITE      = colors.white

    W, H      = A4          # 595.27 x 841.89 pt
    MX        = 40          # horizontal margin
    HEADER_H  = 60
    FOOTER_H  = 36
    CONTENT_W = W - 2 * MX

    generated = datetime.date.today().strftime('%d %b %Y')

    # ── Helper: colour-coded value ────────────────────────────────────────────
    def traffic(value: float, warn: float, danger: float, invert=False) -> colors.Color:
        """Return a colour given warn/danger thresholds. invert=True: lower is worse."""
        if invert:
            return C_DANGER if value < danger else (C_WARNING if value < warn else C_SUCCESS)
        return C_DANGER if value >= danger else (C_WARNING if value >= warn else C_PRIMARY)

    # ── Page template (header + footer drawn directly on canvas) ─────────────
    def on_page(canvas, doc):
        canvas.saveState()

        # Dark header band
        canvas.setFillColor(C_DARK)
        canvas.rect(0, H - HEADER_H, W, HEADER_H, fill=1, stroke=0)

        # Left accent stripe
        canvas.setFillColor(C_PRIMARY)
        canvas.rect(0, H - HEADER_H, 5, HEADER_H, fill=1, stroke=0)

        # Agency name
        canvas.setFillColor(C_WHITE)
        canvas.setFont('Helvetica-Bold', 15)
        canvas.drawString(MX + 8, H - HEADER_H + 35, 'DesignFlow')

        # "PROJECT REPORT" sub-label
        canvas.setFillColor(C_SLATE_400)
        canvas.setFont('Helvetica', 8)
        canvas.drawString(MX + 8, H - HEADER_H + 18, 'PROJECT REPORT')

        # Page number (header, right-aligned)
        canvas.setFillColor(C_SLATE_400)
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(W - MX, H - HEADER_H + 26, f'Page {doc.page}')

        # Footer rule
        canvas.setStrokeColor(C_SLATE_200)
        canvas.setLineWidth(0.5)
        canvas.line(MX, FOOTER_H, W - MX, FOOTER_H)

        # Footer text
        canvas.setFillColor(C_SLATE_400)
        canvas.setFont('Helvetica', 7.5)
        canvas.drawString(MX, FOOTER_H - 13, f'Generated on {generated}  ·  DesignFlow')

        canvas.restoreState()

    buf  = io.BytesIO()
    frame = Frame(
        MX, FOOTER_H + 6,
        CONTENT_W,
        H - HEADER_H - FOOTER_H - 22,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
    )
    doc = BaseDocTemplate(
        buf, pagesize=A4,
        pageTemplates=[PageTemplate(id='main', frames=[frame], onPage=on_page)],
    )

    # ── Paragraph styles ──────────────────────────────────────────────────────
    def ps(name, font='Helvetica', size=9, color=C_SLATE_700,
           bold=False, leading=None, space_before=0, space_after=0, align=TA_LEFT):
        return ParagraphStyle(
            name,
            fontName='Helvetica-Bold' if bold else font,
            fontSize=size,
            textColor=color,
            leading=leading or size * 1.4,
            spaceBefore=space_before,
            spaceAfter=space_after,
            alignment=align,
        )

    S_TITLE   = ps('title',  size=20, bold=True, color=C_SLATE_900, space_after=4)
    S_META    = ps('meta',   size=9,  color=C_SLATE_400, space_after=0)
    S_H2      = ps('h2',     size=10, bold=True, color=C_SLATE_700, space_before=4, space_after=6)
    S_TBL_LBL = ps('tlbl',  size=9,  color=C_SLATE_600)
    S_TBL_VAL = ps('tval',  size=9,  bold=True, color=C_SLATE_900)
    elements: list = []
    elements.append(Spacer(1, 8))

    # ── Project title & meta strip ────────────────────────────────────────────
    elements.append(Paragraph(project.project_name, S_TITLE))

    meta_parts = [project.client.user.full_name, project.status]
    if project.deadline:
        meta_parts.append(f'Due {project.deadline}')
    if project.category:
        meta_parts.append(project.category)
    elements.append(Paragraph('   ·   '.join(meta_parts), S_META))
    elements.append(Spacer(1, 14))
    elements.append(HRFlowable(width='100%', thickness=1, color=C_SLATE_200, spaceAfter=14))

    # ── KPI strip (4 tiles in one table row) ──────────────────────────────────
    util_col  = traffic(budget_util,  warn=80, danger=100)
    scope_col = traffic(scope_creep,  warn=15, danger=30)
    # Green when EHR meets/exceeds target, red when below target, grey when no hours
    ehr_col   = C_SUCCESS if (target_ehr and display_ehr >= target_ehr) else (C_DANGER if display_ehr > 0 else C_SLATE_400)

    def kpi_cell(label: str, value: str, val_color=C_SLATE_900):
        return [
            Paragraph(label, ps(f'kl_{label}', size=7, color=C_SLATE_400, space_after=4)),
            Paragraph(value, ps(f'kv_{label}', size=16, bold=True, color=val_color)),
        ]

    kpi_col_w = (CONTENT_W - 3) / 4
    kpi_table = Table(
        [[
            kpi_cell('BUDGET UTILISATION', f'{budget_util:.0f}%',    util_col),
            kpi_cell('ACTUAL HOURS',       f'{actual_hours:.1f} h',  C_SLATE_900),
            kpi_cell(ehr_label,            f'{display_ehr:.2f} TND' if display_ehr else '—', ehr_col),
            kpi_cell('SCOPE CREEP',        f'{scope_creep:.0f}%',    scope_col),
        ]],
        colWidths=[kpi_col_w] * 4,
    )
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), C_SLATE_50),
        ('BOX',           (0, 0), (-1, -1), 1,   C_SLATE_200),
        ('LINEBEFORE',    (1, 0), (-1, -1), 1,   C_SLATE_200),
        ('LEFTPADDING',   (0, 0), (-1, -1), 14),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 14),
        ('TOPPADDING',    (0, 0), (-1, -1), 13),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 13),
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 22))

    # ── Metrics table (full width) ────────────────────────────────────────────
    elements.append(Paragraph('Metrics', S_H2))

    metrics = [
        ('Budget Amount',         f'{budget_amount:,.2f} TND' if budget_amount else '—'),
        ('Budget Hours',          f'{budget_hours:.1f} h'     if budget_hours  else '—'),
        ('Actual Hours Logged',   f'{actual_hours:.1f} h'),
        ('Estimated Hours',       f'{estimated_h:.1f} h'      if estimated_h   else '—'),
        ('Projected Hourly Rate', f'{display_ehr:.2f} TND'    if display_ehr   else '—'),
        ('Budget Utilisation',    f'{budget_util:.1f}%'),
        ('Scope Creep Index',     f'{scope_creep:.1f}%'),
        ('Revisions',             str(revisions)),
        ('Approvals',             str(approvals)),
        ('Rev : Approval Ratio',  f'{revisions} : {approvals}'),
    ]
    if display_margin is not None:
        metrics.append((margin_label, f'{display_margin:.1f}%'))

    LBL_W = CONTENT_W * 0.50
    VAL_W = CONTENT_W - LBL_W

    metric_rows = [
        [Paragraph(lbl, S_TBL_LBL), Paragraph(val, S_TBL_VAL)]
        for lbl, val in metrics
    ]
    metrics_tbl = Table(metric_rows, colWidths=[LBL_W, VAL_W])
    row_bgs = [
        ('BACKGROUND', (0, i), (-1, i), C_WHITE if i % 2 == 0 else C_SLATE_50)
        for i in range(len(metric_rows))
    ]
    metrics_tbl.setStyle(TableStyle([
        *row_bgs,
        ('BOX',           (0, 0), (-1, -1), 1,   C_SLATE_200),
        ('LINEBELOW',     (0, 0), (-1, -2), 0.5, C_SLATE_100),
        ('LEFTPADDING',   (0, 0), (-1, -1), 12),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 12),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(metrics_tbl)
    elements.append(Spacer(1, 20))

    # ── Build ─────────────────────────────────────────────────────────────────
    doc.build(elements)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# Excel — client profitability + budget data (all projects or one)
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(project_id: int | None = None) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Style constants ───────────────────────────────────────────────────────
    INDIGO_FILL  = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    ALT_FILL     = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    HDR_FONT     = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
    BODY_FONT    = Font(name='Calibri', size=10)
    BOLD_FONT    = Font(bold=True, name='Calibri', size=10)
    CENTER       = Alignment(horizontal='center', vertical='center')
    LEFT         = Alignment(horizontal='left',   vertical='center')
    RIGHT        = Alignment(horizontal='right',  vertical='center')
    THIN         = Side(style='thin', color='E2E8F0')
    BORDER       = Border(bottom=THIN)

    def style_header(ws, headers: list[str]):
        ws.row_dimensions[1].height = 28
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill      = INDIGO_FILL
            c.font      = HDR_FONT
            c.alignment = CENTER

    def style_row(ws, row_idx: int, n_cols: int, alternate: bool):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=row_idx, column=col)
            if alternate:
                c.fill = ALT_FILL
            c.font      = BODY_FONT
            c.alignment = LEFT
            c.border    = BORDER
        ws.row_dimensions[row_idx].height = 20

    def right_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.alignment = RIGHT
        c.font = BOLD_FONT

    def autofit(ws, min_w=12, max_w=40):
        for col_cells in ws.columns:
            length = max(
                len(str(cell.value or ''))
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 3, min_w), max_w)

    def freeze(ws):
        ws.freeze_panes = 'A2'

    # ── Sheet 1: Client Profitability ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Client Profitability'
    headers1 = ['Client', 'Total Revenue (TND)', 'Total Hours', 'Eff. Hourly Rate', 'Revisions', 'Approvals']
    style_header(ws1, headers1)
    freeze(ws1)

    clients = Client.objects.select_related('user').all()

    # Compute ordering key separately so we can sort without annotate fan-out
    client_revenues = {
        row['client_id']: float(row['rev'] or 0)
        for row in Project.objects.values('client_id').annotate(rev=Sum('budget_amount'))
        }
    clients = sorted(clients, key=lambda c: client_revenues.get(c.id, 0), reverse=True)

    for i, c in enumerate(clients, 2):
        projects_qs = Project.objects.filter(client=c)
        revenue = float(projects_qs.aggregate(t=Sum('budget_amount'))['t'] or 0)
        hours   = float(
            TimeLog.objects.filter(task__project__in=projects_qs)
            .aggregate(t=Sum('hours_spent'))['t'] or 0
        )
        revisions = Feedback.objects.filter(project__in=projects_qs, category='Revision').count()
        approvals  = Feedback.objects.filter(project__in=projects_qs, category='Approval').count()
        ehr = round(revenue / hours, 2) if hours > 0 else None

        ws1.cell(row=i, column=1, value=c.user.full_name)
        right_cell(ws1, i, 2, round(revenue, 2))
        right_cell(ws1, i, 3, round(hours,   2))
        right_cell(ws1, i, 4, ehr or '—')
        right_cell(ws1, i, 5, revisions)
        right_cell(ws1, i, 6, approvals)
        style_row(ws1, i, len(headers1), i % 2 == 0)

    autofit(ws1)

    # ── Sheet 2: Budget Data ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Budget Data')
    headers2 = [
        'Project', 'Client', 'Status',
        'Budget Hours', 'Actual Hours', 'Estimated Hours', 'Variance %',
        'Budget Amount (TND)', 'Eff. Hourly Rate', 'Scope Creep %', 'Profit Margin %',
    ]
    style_header(ws2, headers2)
    freeze(ws2)

    projects = Project.objects.select_related('client__user').all()
    if project_id:
        projects = projects.filter(id=project_id)

    for i, p in enumerate(projects, 2):
        task_qs = Task.objects.filter(project=p)
        agg = task_qs.aggregate(
            estimated=Sum('estimated_hours'),
            total=Count('id'),
            unplanned=Count('id', filter=Q(is_unplanned=True)),
        )

        # Query actual hours separately to avoid join duplication
        actual = float(
            TimeLog.objects.filter(task__project=p).aggregate(t=Sum('hours_spent'))['t'] or 0
        )
        estimated = float(agg['estimated'] or 0)
        budget_h  = float(p.budget_hours   or 0)
        budget_a  = float(p.budget_amount  or 0)
        variance  = round((actual - estimated) / estimated * 100, 1) if estimated > 0 else None
        ehr_val   = round(budget_a / actual, 2) if actual > 0 else None
        total_t   = agg['total'] or 0
        scope     = round(agg['unplanned'] / total_t * 100, 1) if total_t > 0 else 0

        weighted_rate = _weighted_designer_rate(TimeLog.objects.filter(task__project=p), actual)
        margin_val = (
            round((ehr_val - weighted_rate) / ehr_val * 100, 1)
            if weighted_rate is not None and ehr_val else None
        )

        row_data = [
            p.project_name, p.client.user.full_name, p.status,
            budget_h, actual, estimated,
            variance if variance is not None else '—',
            round(budget_a, 2),
            ehr_val if ehr_val is not None else '—',
            scope,
            margin_val if margin_val is not None else '—',
        ]
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.alignment = RIGHT if col > 3 else LEFT
            c.font = BOLD_FONT if col > 3 else BODY_FONT

        style_row(ws2, i, len(headers2), i % 2 == 0)

    autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
